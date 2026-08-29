"""
Census Assistant - Field Attendance Module

Daily attendance marking for Enumerators and Supervisors.

Design rules (agreed with the circle office):

* One row per mobile number per calendar day (IST). Re-submitting on the same
  day UPDATES that row — it never creates a second one — so the register can
  never hold duplicate information for the same person on the same day.
* Name / Position / HLB-or-Circle number carry forward automatically from the
  person's most recent record, so a returning user only has to re-take the
  photo and re-confirm their location each day.
* A live device location (latitude/longitude) is mandatory: an attendance mark
  without a location is meaningless for field verification.
* Photo is mandatory the first time a given day's row is created, and optional
  afterwards (the already-stored photo is kept if no new one is sent).
* Admin workflow: PENDING -> APPROVED (photo file is permanently deleted from
  disk, record is locked against further edits) or PENDING -> REJECTED (photo
  is kept, and the user may correct and resubmit, which returns it to PENDING).
"""

import io
import os
import re
import uuid
import logging
from datetime import datetime, timedelta, timezone
from typing import Any, Dict, List, Optional, Tuple

from .database import get_db_connection

logger = logging.getLogger("AttendanceService")

ROOT_DIR = os.path.dirname(os.path.dirname(__file__))
PHOTO_DIR = os.path.join(ROOT_DIR, "attendance_photos")

# India Standard Time. The whole circle operates in one timezone, so the
# "attendance day" is defined in IST regardless of where the server runs
# (Render/Docker containers default to UTC, which would otherwise roll the
# date over at 05:30 IST — right in the middle of a field morning).
IST = timezone(timedelta(hours=5, minutes=30))

VALID_POSITIONS = ("Enumerator", "Supervisor")
VALID_STATUSES = ("PENDING", "APPROVED", "REJECTED")

ALLOWED_PHOTO_EXTENSIONS = {".jpg", ".jpeg", ".png", ".webp"}
MAX_PHOTO_BYTES = 8 * 1024 * 1024  # 8 MB — the client downsizes to ~200-400 KB

MAX_NAME_LEN = 100
MAX_BLOCK_LEN = 30
MAX_REASON_LEN = 500


# --------------------------------------------------------------------------- #
# Schema                                                                       #
# --------------------------------------------------------------------------- #

def init_attendance_schema(cursor) -> None:
    """
    Create the attendance table + indexes. Called from database.init_database()
    inside the existing transaction, so it shares that connection and commit.
    """
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS attendance_records (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            mobile_number TEXT NOT NULL,
            attendance_date TEXT NOT NULL,
            name TEXT NOT NULL,
            position TEXT NOT NULL,
            block_number TEXT NOT NULL,
            latitude REAL NOT NULL,
            longitude REAL NOT NULL,
            accuracy_m REAL,
            photo_filename TEXT,
            photo_deleted INTEGER DEFAULT 0,
            status TEXT DEFAULT 'PENDING',
            reject_reason TEXT,
            reviewed_by TEXT,
            reviewed_at TEXT,
            submitted_at TEXT,
            updated_at TEXT,
            submission_count INTEGER DEFAULT 1,
            user_id TEXT,
            UNIQUE(mobile_number, attendance_date)
        )
    """)
    cursor.execute("CREATE INDEX IF NOT EXISTS idx_attendance_date ON attendance_records(attendance_date)")
    cursor.execute("CREATE INDEX IF NOT EXISTS idx_attendance_mobile ON attendance_records(mobile_number)")
    cursor.execute("CREATE INDEX IF NOT EXISTS idx_attendance_status ON attendance_records(status)")
    cursor.execute("CREATE INDEX IF NOT EXISTS idx_attendance_position ON attendance_records(position)")

    os.makedirs(PHOTO_DIR, exist_ok=True)


# --------------------------------------------------------------------------- #
# Small helpers                                                                #
# --------------------------------------------------------------------------- #

def today_ist() -> str:
    return datetime.now(IST).strftime("%Y-%m-%d")


def now_ist_str() -> str:
    return datetime.now(IST).strftime("%Y-%m-%d %H:%M:%S")


def normalize_mobile(raw: Any) -> Optional[str]:
    """Reduce any input to a bare 10-digit Indian mobile number, or None."""
    digits = "".join(c for c in str(raw or "") if c.isdigit())
    if len(digits) > 10:
        digits = digits[-10:]          # strip 91 / 091 / +91 country prefixes
    if len(digits) != 10 or digits[0] not in "6789":
        return None
    return digits


def _clean_text(raw: Any, max_len: int) -> str:
    text = re.sub(r"\s+", " ", str(raw or "")).strip()
    return text[:max_len]


def _maps_link(lat: Optional[float], lon: Optional[float]) -> str:
    if lat is None or lon is None:
        return ""
    return f"https://www.google.com/maps?q={lat:.6f},{lon:.6f}"


def _row_to_dict(row, include_photo_name: bool = False) -> Dict[str, Any]:
    rec = {
        "id": row["id"],
        "mobile_number": row["mobile_number"],
        "attendance_date": row["attendance_date"],
        "name": row["name"],
        "position": row["position"],
        "block_number": row["block_number"],
        "latitude": row["latitude"],
        "longitude": row["longitude"],
        "accuracy_m": row["accuracy_m"],
        "maps_link": _maps_link(row["latitude"], row["longitude"]),
        "has_photo": bool(row["photo_filename"]) and not row["photo_deleted"],
        "photo_deleted": bool(row["photo_deleted"]),
        "status": row["status"],
        "reject_reason": row["reject_reason"],
        "reviewed_by": row["reviewed_by"],
        "reviewed_at": row["reviewed_at"],
        "submitted_at": row["submitted_at"],
        "updated_at": row["updated_at"],
        "submission_count": row["submission_count"],
        "editable": row["status"] != "APPROVED",
    }
    if include_photo_name:
        rec["photo_filename"] = row["photo_filename"]
    return rec


# --------------------------------------------------------------------------- #
# Photo storage                                                                #
# --------------------------------------------------------------------------- #

def _save_photo(file_storage, mobile: str, date_str: str) -> Tuple[Optional[str], Optional[str]]:
    """
    Persist an uploaded photo under attendance_photos/<date>/.
    Returns (relative_path, error_message).
    """
    filename = (getattr(file_storage, "filename", "") or "").strip()
    ext = os.path.splitext(filename)[1].lower()
    if ext not in ALLOWED_PHOTO_EXTENSIONS:
        # Browsers that hand us a blob with no filename still declare a mimetype.
        mimetype = (getattr(file_storage, "mimetype", "") or "").lower()
        ext = {
            "image/jpeg": ".jpg", "image/jpg": ".jpg",
            "image/png": ".png", "image/webp": ".webp",
        }.get(mimetype, "")
    if ext not in ALLOWED_PHOTO_EXTENSIONS:
        return None, "Photo must be a JPG, PNG or WEBP image."

    blob = file_storage.read()
    if not blob:
        return None, "The uploaded photo was empty."
    if len(blob) > MAX_PHOTO_BYTES:
        return None, "Photo is too large. Please retake it (maximum 8 MB)."

    day_dir = os.path.join(PHOTO_DIR, date_str)
    os.makedirs(day_dir, exist_ok=True)
    stored_name = f"{mobile}_{uuid.uuid4().hex[:10]}{ext}"
    with open(os.path.join(day_dir, stored_name), "wb") as fh:
        fh.write(blob)
    return f"{date_str}/{stored_name}", None


def resolve_photo_path(relative_name: str) -> Optional[str]:
    """
    Map a stored photo_filename to an absolute path, refusing anything that
    escapes PHOTO_DIR (path traversal guard).
    """
    if not relative_name:
        return None
    full = os.path.abspath(os.path.join(PHOTO_DIR, relative_name))
    if not full.startswith(os.path.abspath(PHOTO_DIR) + os.sep):
        return None
    return full if os.path.isfile(full) else None


def _delete_photo(relative_name: str) -> bool:
    path = resolve_photo_path(relative_name)
    if not path:
        return False
    try:
        os.remove(path)
        return True
    except OSError as exc:
        logger.warning(f"Could not delete attendance photo {relative_name}: {exc}")
        return False


# --------------------------------------------------------------------------- #
# Reads                                                                        #
# --------------------------------------------------------------------------- #

def get_attendance(mobile: str, date_str: Optional[str] = None) -> Optional[Dict[str, Any]]:
    """Fetch one person's record for a given day (defaults to today, IST)."""
    clean = normalize_mobile(mobile)
    if not clean:
        return None
    date_str = date_str or today_ist()

    conn = get_db_connection()
    try:
        row = conn.execute(
            "SELECT * FROM attendance_records WHERE mobile_number = ? AND attendance_date = ?",
            (clean, date_str),
        ).fetchone()
    finally:
        conn.close()
    return _row_to_dict(row) if row else None


def get_carry_forward_profile(mobile: str) -> Optional[Dict[str, Any]]:
    """
    Name / position / block from this person's most recent record, used to
    prefill the form so a returning user only re-takes photo + location.
    """
    clean = normalize_mobile(mobile)
    if not clean:
        return None

    conn = get_db_connection()
    try:
        row = conn.execute("""
            SELECT name, position, block_number, attendance_date
            FROM attendance_records
            WHERE mobile_number = ?
            ORDER BY attendance_date DESC, id DESC
            LIMIT 1
        """, (clean,)).fetchone()
    finally:
        conn.close()

    if not row:
        return None
    return {
        "name": row["name"],
        "position": row["position"],
        "block_number": row["block_number"],
        "from_date": row["attendance_date"],
    }


def _build_filter_sql(filters: Dict[str, Any]) -> Tuple[str, List[Any]]:
    clauses, params = [], []

    status = (filters.get("status") or "").upper()
    if status in VALID_STATUSES:
        clauses.append("status = ?")
        params.append(status)

    position = (filters.get("position") or "").strip().title()
    if position in VALID_POSITIONS:
        clauses.append("position = ?")
        params.append(position)

    date_from = (filters.get("date_from") or "").strip()
    if re.fullmatch(r"\d{4}-\d{2}-\d{2}", date_from):
        clauses.append("attendance_date >= ?")
        params.append(date_from)

    date_to = (filters.get("date_to") or "").strip()
    if re.fullmatch(r"\d{4}-\d{2}-\d{2}", date_to):
        clauses.append("attendance_date <= ?")
        params.append(date_to)

    query = (filters.get("q") or "").strip()
    if query:
        like = f"%{query}%"
        clauses.append("(name LIKE ? OR mobile_number LIKE ? OR block_number LIKE ?)")
        params.extend([like, like, like])

    where = (" WHERE " + " AND ".join(clauses)) if clauses else ""
    return where, params


def list_attendance(filters: Dict[str, Any], limit: int = 200, offset: int = 0) -> Dict[str, Any]:
    """Admin listing with the filter set exposed in the console."""
    where, params = _build_filter_sql(filters)

    conn = get_db_connection()
    try:
        total = conn.execute(
            f"SELECT COUNT(*) AS n FROM attendance_records{where}", params
        ).fetchone()["n"]

        rows = conn.execute(
            f"""SELECT * FROM attendance_records{where}
                ORDER BY attendance_date DESC, updated_at DESC, id DESC
                LIMIT ? OFFSET ?""",
            (*params, limit, offset),
        ).fetchall()

        counts = {r["status"]: r["n"] for r in conn.execute(
            "SELECT status, COUNT(*) AS n FROM attendance_records GROUP BY status"
        ).fetchall()}
    finally:
        conn.close()

    return {
        "records": [_row_to_dict(r) for r in rows],
        "total": total,
        "limit": limit,
        "offset": offset,
        "summary": {
            "pending": counts.get("PENDING", 0),
            "approved": counts.get("APPROVED", 0),
            "rejected": counts.get("REJECTED", 0),
            "all": sum(counts.values()),
        },
    }


# --------------------------------------------------------------------------- #
# Writes                                                                       #
# --------------------------------------------------------------------------- #

def submit_attendance(form: Dict[str, Any], photo_file=None, user_id: Optional[str] = None) -> Tuple[Dict[str, Any], int]:
    """
    Create or update today's attendance row for one mobile number.
    Returns (response_body, http_status).
    """
    mobile = normalize_mobile(form.get("mobile_number"))
    if not mobile:
        return {"success": False, "error": "Enter a valid 10-digit mobile number."}, 400

    name = _clean_text(form.get("name"), MAX_NAME_LEN)
    if len(name) < 2:
        return {"success": False, "error": "Enter your full name."}, 400

    position = _clean_text(form.get("position"), 20).title()
    if position not in VALID_POSITIONS:
        return {"success": False, "error": "Select a position: Enumerator or Supervisor."}, 400

    block_number = _clean_text(form.get("block_number"), MAX_BLOCK_LEN)
    if not block_number:
        label = "HLB number" if position == "Enumerator" else "Supervisory Circle number"
        return {"success": False, "error": f"Enter your {label}."}, 400

    try:
        latitude = float(form.get("latitude"))
        longitude = float(form.get("longitude"))
    except (TypeError, ValueError):
        return {"success": False,
                "error": "Location not captured. Tap 'Capture My Location' and allow location access."}, 400

    if not (-90.0 <= latitude <= 90.0 and -180.0 <= longitude <= 180.0):
        return {"success": False, "error": "The captured location is not valid. Please capture it again."}, 400

    try:
        accuracy = float(form.get("accuracy_m")) if form.get("accuracy_m") not in (None, "") else None
    except (TypeError, ValueError):
        accuracy = None

    date_str = today_ist()
    now = now_ist_str()
    has_upload = photo_file is not None and getattr(photo_file, "filename", None) is not None

    conn = get_db_connection()
    try:
        conn.execute("BEGIN IMMEDIATE")
        existing = conn.execute(
            "SELECT * FROM attendance_records WHERE mobile_number = ? AND attendance_date = ?",
            (mobile, date_str),
        ).fetchone()

        # Approved records are final — this is what stops an approved entry
        # being quietly rewritten after the fact.
        if existing and existing["status"] == "APPROVED":
            conn.rollback()
            return {
                "success": False,
                "locked": True,
                "error": "Today's attendance has already been approved and can no longer be edited. "
                         "Contact the Technical Assistant if a correction is needed.",
                "record": _row_to_dict(existing),
            }, 409

        if not existing and not has_upload:
            conn.rollback()
            return {"success": False, "error": "A photo is required for today's attendance."}, 400

        new_photo_rel = None
        if has_upload:
            new_photo_rel, err = _save_photo(photo_file, mobile, date_str)
            if err:
                conn.rollback()
                return {"success": False, "error": err}, 400

        if existing:
            # Replacing the photo: write the new one first, then drop the old
            # file, so a failed write never leaves the record photo-less.
            old_photo = existing["photo_filename"]
            photo_rel = new_photo_rel or old_photo
            conn.execute("""
                UPDATE attendance_records
                SET name = ?, position = ?, block_number = ?,
                    latitude = ?, longitude = ?, accuracy_m = ?,
                    photo_filename = ?, photo_deleted = 0,
                    status = 'PENDING', reject_reason = NULL,
                    reviewed_by = NULL, reviewed_at = NULL,
                    updated_at = ?, submission_count = submission_count + 1,
                    user_id = COALESCE(?, user_id)
                WHERE id = ?
            """, (name, position, block_number, latitude, longitude, accuracy,
                  photo_rel, now, user_id, existing["id"]))
            conn.commit()
            if new_photo_rel and old_photo and old_photo != new_photo_rel:
                _delete_photo(old_photo)
            record_id, created = existing["id"], False
        else:
            cur = conn.execute("""
                INSERT INTO attendance_records
                    (mobile_number, attendance_date, name, position, block_number,
                     latitude, longitude, accuracy_m, photo_filename, photo_deleted,
                     status, submitted_at, updated_at, submission_count, user_id)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, 0, 'PENDING', ?, ?, 1, ?)
            """, (mobile, date_str, name, position, block_number,
                  latitude, longitude, accuracy, new_photo_rel, now, now, user_id))
            conn.commit()
            record_id, created = cur.lastrowid, True

        row = conn.execute("SELECT * FROM attendance_records WHERE id = ?", (record_id,)).fetchone()
    except Exception as exc:
        conn.rollback()
        logger.exception("Attendance submission failed")
        return {"success": False, "error": f"Could not save attendance: {exc}"}, 500
    finally:
        conn.close()

    return {
        "success": True,
        "created": created,
        "message": "Attendance submitted." if created else "Attendance updated.",
        "record": _row_to_dict(row),
    }, 200


def review_attendance(record_id: int, action: str, reviewer: str, reason: str = "") -> Tuple[Dict[str, Any], int]:
    """
    Admin approve / reject.

    APPROVE permanently deletes the photo file from disk — the photo exists only
    to let the admin verify the person, and is not part of the retained record.
    REJECT keeps the photo so the user can see what was wrong and resubmit.
    """
    action = (action or "").lower()
    if action not in ("approve", "reject"):
        return {"success": False, "error": "Unknown review action."}, 400

    reason = _clean_text(reason, MAX_REASON_LEN)
    if action == "reject" and not reason:
        return {"success": False, "error": "Give a reason so the user knows what to correct."}, 400

    conn = get_db_connection()
    try:
        row = conn.execute("SELECT * FROM attendance_records WHERE id = ?", (record_id,)).fetchone()
        if not row:
            return {"success": False, "error": "Attendance record not found."}, 404

        now = now_ist_str()
        if action == "approve":
            removed = _delete_photo(row["photo_filename"]) if row["photo_filename"] else False
            conn.execute("""
                UPDATE attendance_records
                SET status = 'APPROVED', reject_reason = NULL,
                    photo_filename = NULL, photo_deleted = 1,
                    reviewed_by = ?, reviewed_at = ?, updated_at = ?
                WHERE id = ?
            """, (reviewer, now, now, record_id))
            message = "Approved. Photo deleted from the server."
            if row["photo_filename"] and not removed:
                message = "Approved. Photo file was already missing from disk."
        else:
            conn.execute("""
                UPDATE attendance_records
                SET status = 'REJECTED', reject_reason = ?,
                    reviewed_by = ?, reviewed_at = ?, updated_at = ?
                WHERE id = ?
            """, (reason, reviewer, now, now, record_id))
            message = "Rejected. The user can correct and resubmit."

        conn.commit()
        updated = conn.execute("SELECT * FROM attendance_records WHERE id = ?", (record_id,)).fetchone()
    finally:
        conn.close()

    return {"success": True, "message": message, "record": _row_to_dict(updated)}, 200


def delete_attendance(record_id: int) -> Tuple[Dict[str, Any], int]:
    """Remove a record entirely (and its photo). Admin-only escape hatch."""
    conn = get_db_connection()
    try:
        row = conn.execute("SELECT * FROM attendance_records WHERE id = ?", (record_id,)).fetchone()
        if not row:
            return {"success": False, "error": "Attendance record not found."}, 404
        if row["photo_filename"]:
            _delete_photo(row["photo_filename"])
        conn.execute("DELETE FROM attendance_records WHERE id = ?", (record_id,))
        conn.commit()
    finally:
        conn.close()
    return {"success": True, "message": "Attendance record deleted."}, 200


def get_photo_for_record(record_id: int) -> Optional[str]:
    """Absolute path of a record's photo, or None if approved/missing."""
    conn = get_db_connection()
    try:
        row = conn.execute(
            "SELECT photo_filename FROM attendance_records WHERE id = ?", (record_id,)
        ).fetchone()
    finally:
        conn.close()
    if not row or not row["photo_filename"]:
        return None
    return resolve_photo_path(row["photo_filename"])


def purge_orphaned_photos() -> int:
    """
    Housekeeping: delete photo files on disk that no live record points at
    (e.g. left behind by an interrupted replace). Returns the count removed.
    """
    conn = get_db_connection()
    try:
        keep = {
            r["photo_filename"] for r in conn.execute(
                "SELECT photo_filename FROM attendance_records WHERE photo_filename IS NOT NULL"
            ).fetchall()
        }
    finally:
        conn.close()

    removed = 0
    if not os.path.isdir(PHOTO_DIR):
        return 0
    for day in os.listdir(PHOTO_DIR):
        day_path = os.path.join(PHOTO_DIR, day)
        if not os.path.isdir(day_path):
            continue
        for fname in os.listdir(day_path):
            if f"{day}/{fname}" not in keep:
                try:
                    os.remove(os.path.join(day_path, fname))
                    removed += 1
                except OSError:
                    pass
        if not os.listdir(day_path):
            try:
                os.rmdir(day_path)
            except OSError:
                pass
    return removed


# --------------------------------------------------------------------------- #
# Excel export                                                                 #
# --------------------------------------------------------------------------- #

EXPORT_COLUMNS = [
    ("Date", 12),
    ("Name", 26),
    ("Mobile Number", 15),
    ("Position", 13),
    ("HLB / Circle No", 17),
    ("Latitude", 12),
    ("Longitude", 12),
    ("Accuracy (m)", 13),
    ("Location Link", 40),
    ("Status", 11),
    ("Rejection Reason", 30),
    ("Submissions", 12),
    ("First Submitted (IST)", 21),
    ("Last Updated (IST)", 21),
    ("Reviewed By", 18),
    ("Reviewed At (IST)", 21),
    ("Photo", 22),
]


def build_attendance_workbook(filters: Dict[str, Any]) -> Tuple[io.BytesIO, str]:
    """
    Build one .xlsx holding every attendance row matching *filters* — all
    users, all days, in a single sheet. Returns (buffer, suggested_filename).
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    where, params = _build_filter_sql(filters)
    conn = get_db_connection()
    try:
        rows = conn.execute(
            f"""SELECT * FROM attendance_records{where}
                ORDER BY attendance_date DESC, position, name""",
            params,
        ).fetchall()
    finally:
        conn.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance"

    header_fill = PatternFill("solid", fgColor="1A4E8A")
    header_font = Font(bold=True, color="FFFFFF", size=10)
    thin = Side(style="thin", color="D5DCE4")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.append([c[0] for c in EXPORT_COLUMNS])
    for idx, (title, width) in enumerate(EXPORT_COLUMNS, start=1):
        cell = ws.cell(row=1, column=idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
        ws.column_dimensions[get_column_letter(idx)].width = width
    ws.row_dimensions[1].height = 28

    status_colors = {"APPROVED": "1B7F3B", "REJECTED": "B3261E", "PENDING": "8A6100"}

    for row in rows:
        if row["photo_deleted"]:
            photo_note = "Deleted after approval"
        elif row["photo_filename"]:
            photo_note = "Awaiting review"
        else:
            photo_note = "—"

        ws.append([
            row["attendance_date"],
            row["name"],
            row["mobile_number"],
            row["position"],
            row["block_number"],
            row["latitude"],
            row["longitude"],
            row["accuracy_m"],
            _maps_link(row["latitude"], row["longitude"]),
            row["status"].title(),
            row["reject_reason"] or "",
            row["submission_count"],
            row["submitted_at"] or "",
            row["updated_at"] or "",
            row["reviewed_by"] or "",
            row["reviewed_at"] or "",
            photo_note,
        ])

        r = ws.max_row
        for c in range(1, len(EXPORT_COLUMNS) + 1):
            ws.cell(row=r, column=c).border = border
            ws.cell(row=r, column=c).alignment = Alignment(vertical="center")
        ws.cell(row=r, column=6).number_format = "0.000000"
        ws.cell(row=r, column=7).number_format = "0.000000"
        link_cell = ws.cell(row=r, column=9)
        if link_cell.value:
            link_cell.hyperlink = link_cell.value
            link_cell.font = Font(color="1155CC", underline="single", size=10)
        status_cell = ws.cell(row=r, column=10)
        status_cell.font = Font(bold=True, color=status_colors.get(row["status"], "000000"), size=10)

    ws.freeze_panes = "A2"
    if ws.max_row >= 1:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(EXPORT_COLUMNS))}{max(ws.max_row, 1)}"

    if not rows:
        ws.cell(row=2, column=1, value="No attendance records match the selected filters.")

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    span = ""
    if filters.get("date_from") or filters.get("date_to"):
        span = f"_{filters.get('date_from') or 'start'}_to_{filters.get('date_to') or 'today'}"
    status_tag = f"_{filters['status'].title()}" if (filters.get("status") or "").upper() in VALID_STATUSES else ""
    filename = f"Census_Attendance{status_tag}{span}_{datetime.now(IST).strftime('%Y%m%d_%H%M')}.xlsx"
    return buffer, filename
