"""
Census Assistant - Ingestion Pipeline
Parses Excel source sheets (Functionaries & HLB Allocations) and PDF manuals into structured tables and FTS search indices.
"""

import os
import re
import logging
from datetime import datetime
from typing import List, Optional, Tuple
import openpyxl
import pypdf
from .database import get_db_connection, init_database

logger = logging.getLogger("IngestionEngine")
logging.basicConfig(level=logging.INFO)

ROOT_DIR = os.path.dirname(os.path.dirname(__file__))

def excel_data_type(filename: str) -> str:
    """
    Classify an uploaded Excel file by filename keyword so upload, delete,
    and re-sync all agree on which table it backs. Shared by upload_excel()
    and remove_source_file_data() in main.py so the routing logic can never
    drift out of sync between the two.
    """
    lower = (filename or "").lower()
    if "description" in lower:
        return "hlb_description"
    if "user" in lower:
        return "users"
    return "hlb_allocation"


def find_latest_source(data_type: str) -> Optional[str]:
    """
    Locate the newest spreadsheet in ROOT_DIR that backs *data_type*.

    Source spreadsheets do not keep a stable filename in practice. Each time
    the circle office re-downloads one, the browser appends a counter --
    "All_Users (3).xlsx" -- and secure_filename() rewrites that to
    "All_Users_3.xlsx" on upload. The previous code looked for the exact
    original names, so after the first re-upload a Force Sync silently found
    nothing and reported zero rows for every spreadsheet.

    Files are matched on excel_data_type(), the same classifier the upload and
    delete paths already use, so all three agree on which table a file backs.
    When several candidates exist (old copies left behind by successive
    uploads), the most recently modified one wins.
    """
    candidates = []
    for fname in os.listdir(ROOT_DIR):
        if os.path.splitext(fname)[1].lower() not in (".xlsx", ".xls"):
            continue
        if fname.startswith("~$"):          # Excel lock files
            continue
        if excel_data_type(fname) != data_type:
            continue
        full = os.path.join(ROOT_DIR, fname)
        try:
            candidates.append((os.path.getmtime(full), full))
        except OSError:
            continue

    if not candidates:
        return None

    candidates.sort(reverse=True)
    chosen = candidates[0][1]
    if len(candidates) > 1:
        others = ", ".join(os.path.basename(p) for _, p in candidates[1:])
        logger.info(
            f"{data_type}: {len(candidates)} candidate files present; using the newest "
            f"({os.path.basename(chosen)}). Older copies ignored: {others}"
        )
    return chosen


def remove_source_file_data(filename: str) -> dict:
    """
    Delete the DB rows that came from a specific uploaded source file, used
    when an admin deletes that file from the Admin Panel. Previously,
    deleting a file just removed it from disk and re-ran full ingestion —
    which silently skipped re-processing (since the file was now gone)
    WITHOUT ever clearing the rows it had originally inserted, so "deleted"
    data kept showing up everywhere in the app. This targets exactly the
    rows that file is responsible for.
    """
    ext = os.path.splitext(filename)[1].lower()
    conn = get_db_connection()
    cursor = conn.cursor()
    removed = {}

    if ext == ".pdf":
        cursor.execute("SELECT COUNT(*) FROM manual_chunks WHERE source_file = ?", (filename,))
        removed["manual_chunks"] = cursor.fetchone()[0]
        cursor.execute("DELETE FROM manual_chunks WHERE source_file = ?", (filename,))
        try:
            cursor.execute("DELETE FROM manual_chunks_fts WHERE source_file = ?", (filename,))
        except Exception:
            pass
    else:
        data_type = excel_data_type(filename)
        if data_type == "hlb_description":
            cursor.execute("SELECT COUNT(*) FROM hlb_descriptions")
            removed["hlb_descriptions"] = cursor.fetchone()[0]
            cursor.execute("DELETE FROM hlb_descriptions")
        elif data_type == "users":
            cursor.execute("SELECT COUNT(*) FROM functionaries")
            removed["functionaries"] = cursor.fetchone()[0]
            cursor.execute("DELETE FROM functionaries")
            try:
                cursor.execute("DELETE FROM functionaries_fts")
            except Exception:
                pass
        else:
            cursor.execute("SELECT COUNT(*) FROM hlb_allocations")
            removed["hlb_allocations"] = cursor.fetchone()[0]
            cursor.execute("DELETE FROM hlb_allocations")
            try:
                cursor.execute("DELETE FROM hlb_allocations_fts")
            except Exception:
                pass

    conn.commit()
    conn.close()
    logger.info(f"Removed data sourced from {filename}: {removed}")
    return removed


def ingest_all_users(filepath: str = None) -> int:
    """Ingest All_Users.xlsx into functionaries table."""
    if not filepath:
        # Resolved by content type, not by exact filename -- see find_latest_source.
        filepath = find_latest_source("users") or os.path.join(ROOT_DIR, "All_Users.xlsx")
    if not os.path.exists(filepath):
        logger.warning(f"File not found: {filepath}")
        return 0

    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active

    conn = get_db_connection()
    cursor = conn.cursor()

    cursor.execute("DELETE FROM functionaries")
    try:
        cursor.execute("DELETE FROM functionaries_fts")
    except Exception:
        pass

    inserted = 0
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row or not any(row):
            continue
        sno = row[0] if len(row) > 0 else None
        user_id = str(row[1]).strip() if len(row) > 1 and row[1] is not None else ""
        if not user_id:
            continue
        func_type = str(row[2]).strip() if len(row) > 2 and row[2] is not None else ""
        name = str(row[3]).strip() if len(row) > 3 and row[3] is not None else ""
        mobile = str(row[4]).strip().replace(".0", "") if len(row) > 4 and row[4] is not None else ""
        state_ut = str(row[5]).strip() if len(row) > 5 and row[5] is not None else ""
        district = str(row[6]).strip() if len(row) > 6 and row[6] is not None else ""
        sub_district = str(row[7]).strip() if len(row) > 7 and row[7] is not None else ""
        village_town = str(row[8]).strip() if len(row) > 8 and row[8] is not None else ""
        status = str(row[9]).strip() if len(row) > 9 and row[9] is not None else "ACTIVE"

        cursor.execute("""
            INSERT OR REPLACE INTO functionaries 
            (sno, user_id, functionary_type, name, mobile_number, state_ut, district, sub_district, village_town, status)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (sno, user_id, func_type, name, mobile, state_ut, district, sub_district, village_town, status))

        row_id = cursor.lastrowid
        try:
            cursor.execute("""
                INSERT INTO functionaries_fts (rowid, user_id, functionary_type, name, mobile_number, district, sub_district, village_town)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """, (row_id, user_id, func_type, name, mobile, district, sub_district, village_town))
        except Exception as e:
            logger.debug(f"FTS insert note: {e}")

        inserted += 1

    conn.commit()
    conn.close()
    logger.info(f"Ingested {inserted} functionaries from {filepath}")
    return inserted


def ingest_hlb_allocation(filepath: str = None) -> int:
    """Ingest HLB Allocation (2).xlsx into hlb_allocations table."""
    if not filepath:
        filepath = find_latest_source("hlb_allocation") or os.path.join(ROOT_DIR, "HLB Allocation (2).xlsx")
    if not os.path.exists(filepath):
        logger.warning(f"File not found: {filepath}")
        return 0

    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active

    conn = get_db_connection()
    cursor = conn.cursor()

    cursor.execute("DELETE FROM hlb_allocations")
    try:
        cursor.execute("DELETE FROM hlb_allocations_fts")
    except Exception:
        pass

    inserted = 0
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row or not any(row):
            continue
        circle_no = str(row[0]).strip() if len(row) > 0 and row[0] is not None else ""
        hlb_no = str(row[1]).strip() if len(row) > 1 and row[1] is not None else ""
        if not hlb_no and not circle_no:
            continue
        sup_name = str(row[2]).strip() if len(row) > 2 and row[2] is not None else ""
        enum_name = str(row[3]).strip() if len(row) > 3 and row[3] is not None else ""
        enum_user_id = str(row[4]).strip() if len(row) > 4 and row[4] is not None else ""
        allot_date = str(row[5]).strip() if len(row) > 5 and row[5] is not None else ""

        cursor.execute("""
            INSERT OR REPLACE INTO hlb_allocations 
            (supervisory_circle_no, hlb_no, supervisor_name, enumerator_name, enumerator_user_id, allotment_date)
            VALUES (?, ?, ?, ?, ?, ?)
        """, (circle_no, hlb_no, sup_name, enum_name, enum_user_id, allot_date))

        row_id = cursor.lastrowid
        try:
            cursor.execute("""
                INSERT INTO hlb_allocations_fts (rowid, supervisory_circle_no, hlb_no, supervisor_name, enumerator_name, enumerator_user_id)
                VALUES (?, ?, ?, ?, ?, ?)
            """, (row_id, circle_no, hlb_no, sup_name, enum_name, enum_user_id))
        except Exception:
            pass

        inserted += 1

    conn.commit()
    conn.close()
    logger.info(f"Ingested {inserted} HLB allocations from {filepath}")
    return inserted


def ingest_hlb_description(filepath: str = None) -> int:
    """
    Ingest 'HLB Description.xlsx' (columns: HLB No., Village/Ward name and code,
    Landmark, HLB Description) into hlb_descriptions.
    """
    if not filepath:
        filepath = find_latest_source("hlb_description") or os.path.join(ROOT_DIR, "HLB Description.xlsx")
    if not os.path.exists(filepath):
        logger.warning(f"File not found: {filepath}")
        return 0

    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active

    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("DELETE FROM hlb_descriptions")

    inserted = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not any(row):
            continue
        hlb_no_raw = row[0]
        if hlb_no_raw is None:
            continue
        hlb_no = str(hlb_no_raw).strip()
        if not hlb_no:
            continue
        digits = re.sub(r'\D', '', hlb_no)
        hlb_no_norm = str(int(digits)) if digits else hlb_no

        village_ward = str(row[1]).strip() if len(row) > 1 and row[1] is not None else ""
        landmark = str(row[2]).strip() if len(row) > 2 and row[2] is not None else ""
        description = str(row[3]).strip() if len(row) > 3 and row[3] is not None else ""

        cursor.execute("""
            INSERT OR REPLACE INTO hlb_descriptions (hlb_no, village_ward_name, landmark, boundary_description)
            VALUES (?, ?, ?, ?)
        """, (hlb_no_norm, village_ward, landmark, description))
        inserted += 1

    conn.commit()
    conn.close()
    logger.info(f"Ingested {inserted} HLB descriptions from {filepath}")
    return inserted


# Nicer known titles for the two manuals shipped with the app. Any other
# PDF an admin uploads still gets indexed (see _ingest_single_pdf below) —
# it just falls back to a title derived from its filename.
KNOWN_PDF_TITLES = {
    "FAQ (E & S).c58cf9c49a6df89a94b3 (1).pdf": "Census 2027 FAQ for Enumerators and Supervisors",
    "HLO_Manual_English.pdf": "House Listing Operations (HLO) Instruction Manual",
}


def _ingest_single_pdf(cursor, filepath: str) -> int:
    """
    Parse one PDF into manual_chunks. Only ever touches rows tagged with
    THIS file's name (source_file), so processing/re-processing one manual
    never disturbs chunks belonging to any other manual — which is what
    lets multiple PDFs coexist and get deleted independently.
    """
    filename = os.path.basename(filepath)
    doc_title = KNOWN_PDF_TITLES.get(filename) or (
        os.path.splitext(filename)[0].replace("_", " ").replace("-", " ").strip() or filename
    )

    cursor.execute("DELETE FROM manual_chunks WHERE source_file = ?", (filename,))
    try:
        cursor.execute("DELETE FROM manual_chunks_fts WHERE source_file = ?", (filename,))
    except Exception:
        pass

    if not os.path.exists(filepath):
        logger.warning(f"PDF file not found: {filepath}")
        return 0

    logger.info(f"Parsing PDF: {filename}...")
    chunk_count = 0
    pages_with_text = 0
    boilerplate_pages = 0
    try:
        reader = pypdf.PdfReader(filepath)
        num_pages = len(reader.pages)
        for page_num, page in enumerate(reader.pages, start=1):
            try:
                text = page.extract_text() or ""
            except Exception as e:
                logger.warning(f"Error reading page {page_num} in {filename}: {e}")
                continue

            if not text.strip():
                continue

            # A scanned manual has no text layer, but it is rarely completely
            # empty: prepress software leaves a job ticket on every page
            # ("... Instruction Manual ENGLISH AU-136PAGE DGT.job  Sig13
            # SideA"). Indexing those produced one meaningless chunk per page
            # that could match a search and be cited as if it were guidance.
            # Anything this short, or that looks like press furniture, is not
            # manual content.
            if _is_pdf_boilerplate(text):
                boilerplate_pages += 1
                continue

            pages_with_text += 1

            paragraphs = [p.strip() for p in text.split("\n\n") if p.strip()]
            if not paragraphs:
                paragraphs = [text.strip()]

            current_chunk = ""
            current_header = ""

            for p in paragraphs:
                header_match = re.match(r'^(Q\.\s*\d+|Question\s*\d+|Chapter\s*\d+|\d+\.\d+|\b[A-Z\s]{4,}\b)', p)
                if header_match:
                    current_header = header_match.group(0)

                # Keep chunk within 800 chars, but if a question starts, keep it intact
                if len(current_chunk) + len(p) < 800 and not (current_chunk and header_match):
                    current_chunk += ("\n\n" + p) if current_chunk else p
                else:
                    if current_chunk.strip():
                        cursor.execute("""
                            INSERT INTO manual_chunks (source_file, doc_title, page_number, section_header, chunk_text)
                            VALUES (?, ?, ?, ?, ?)
                        """, (filename, doc_title, page_num, current_header, current_chunk.strip()))
                        row_id = cursor.lastrowid
                        try:
                            cursor.execute("""
                                INSERT INTO manual_chunks_fts (rowid, source_file, doc_title, page_number, section_header, chunk_text)
                                VALUES (?, ?, ?, ?, ?, ?)
                            """, (row_id, filename, doc_title, str(page_num), current_header, current_chunk.strip()))
                        except Exception:
                            pass
                        chunk_count += 1
                    current_chunk = p
                    if header_match:
                        current_header = header_match.group(0)

            if current_chunk.strip():
                cursor.execute("""
                    INSERT INTO manual_chunks (source_file, doc_title, page_number, section_header, chunk_text)
                    VALUES (?, ?, ?, ?, ?)
                """, (filename, doc_title, page_num, current_header, current_chunk.strip()))
                row_id = cursor.lastrowid
                try:
                    cursor.execute("""
                        INSERT INTO manual_chunks_fts (rowid, source_file, doc_title, page_number, section_header, chunk_text)
                        VALUES (?, ?, ?, ?, ?, ?)
                    """, (row_id, filename, doc_title, str(page_num), current_header, current_chunk.strip()))
                except Exception:
                    pass
                chunk_count += 1

        if chunk_count == 0 and boilerplate_pages:
            logger.warning(
                f"{filename}: no readable text found across {num_pages} page(s) — "
                f"{boilerplate_pages} page(s) contained only printer boilerplate. "
                "This PDF is almost certainly a scan of printed pages, so it has no text "
                "layer to index. Run it through OCR and upload the resulting .txt instead."
            )
        else:
            logger.info(
                f"Parsed {chunk_count} chunks from {filename} "
                f"({pages_with_text}/{num_pages} pages had usable text)"
            )
    except Exception as e:
        logger.error(f"Error reading PDF {filename}: {e}")

    return chunk_count


# Printer/prepress furniture that appears on every page of a scanned document.
_BOILERPLATE_HINTS = ("sideA", "sidea", "sideb", ".job", "signature", "sig")


def _is_pdf_boilerplate(text: str) -> bool:
    """
    True when a page's extracted text is press furniture rather than content.

    Scanned PDFs typically yield the same one-line prepress job ticket on
    every page. Left unchecked, each of those became an indexed "manual"
    chunk that the assistant could retrieve and cite.
    """
    stripped = " ".join(text.split())
    if len(stripped) < 120:
        lowered = stripped.lower()
        if any(h in lowered for h in _BOILERPLATE_HINTS):
            return True
        # Too little text to be a page of a manual either way.
        if len(stripped) < 80:
            return True
    return False


def ingest_text_manual(filepath: str) -> int:
    """
    Ingest a plain-text or Markdown manual into manual_chunks.

    This is the route for a manual whose PDF is a scan: OCR it, then upload
    the resulting .txt. The assistant then has real, searchable guidance text
    instead of nothing.

    Page numbers are preserved when the file carries markers of the form

        ===== PAGE 42 =====

    (which is what the OCR output uses) so citations can still say
    "Page 42". Without markers the file is split into sequential blocks and
    numbered from 1, so a citation still points somewhere useful.
    """
    from .database import get_db_connection

    filename = os.path.basename(filepath)
    doc_title = KNOWN_PDF_TITLES.get(filename) or (
        os.path.splitext(filename)[0].replace("_", " ").replace("-", " ").strip() or filename
    )

    if not os.path.exists(filepath):
        logger.warning(f"Text manual not found: {filepath}")
        return 0

    with open(filepath, "r", encoding="utf-8", errors="replace") as fh:
        raw = fh.read()

    conn = get_db_connection()
    cursor = conn.cursor()

    # Replace only this file's rows, exactly like the PDF path, so manuals
    # stay independently re-uploadable and deletable.
    cursor.execute("DELETE FROM manual_chunks WHERE source_file = ?", (filename,))
    try:
        cursor.execute("DELETE FROM manual_chunks_fts WHERE source_file = ?", (filename,))
    except Exception:
        pass

    # Split on page markers if present; otherwise treat the whole file as one
    # stream that gets sliced into numbered blocks below.
    parts = re.split(r'^\s*=+\s*PAGE\s+(\d+)\s*=+\s*$', raw, flags=re.MULTILINE)
    pages: List[Tuple[int, str]] = []
    if len(parts) > 1:
        # re.split with one capture group yields [pre, num, body, num, body...]
        for i in range(1, len(parts) - 1, 2):
            pages.append((int(parts[i]), parts[i + 1]))
    else:
        blocks = [b for b in re.split(r'\n\s*\n\s*\n+', raw) if b.strip()]
        pages = list(enumerate(blocks, start=1)) if blocks else [(1, raw)]

    chunk_count = 0
    for page_num, body in pages:
        body = body.strip()
        if not body:
            continue

        paragraphs = [p.strip() for p in re.split(r'\n\s*\n', body) if p.strip()] or [body]
        current_chunk = ""
        current_header = ""

        def _flush(chunk: str, header: str) -> int:
            if not chunk.strip():
                return 0
            cursor.execute("""
                INSERT INTO manual_chunks (source_file, doc_title, page_number, section_header, chunk_text)
                VALUES (?, ?, ?, ?, ?)
            """, (filename, doc_title, page_num, header, chunk.strip()))
            row_id = cursor.lastrowid
            try:
                cursor.execute("""
                    INSERT INTO manual_chunks_fts (rowid, source_file, doc_title, page_number, section_header, chunk_text)
                    VALUES (?, ?, ?, ?, ?, ?)
                """, (row_id, filename, doc_title, str(page_num), header, chunk.strip()))
            except Exception:
                pass
            return 1

        for p in paragraphs:
            header_match = re.match(
                r'^(Q\.\s*\d+|Question\s*\d+|Chapter\s*\d+|\d+\.\d+|#{1,4}\s*\S.*|\b[A-Z][A-Z\s/&-]{5,}\b)', p
            )
            if header_match and not current_chunk:
                current_header = header_match.group(0).lstrip("#").strip()

            if len(current_chunk) + len(p) < 800 and not (current_chunk and header_match):
                current_chunk += ("\n\n" + p) if current_chunk else p
            else:
                chunk_count += _flush(current_chunk, current_header)
                current_chunk = p
                if header_match:
                    current_header = header_match.group(0).lstrip("#").strip()

        chunk_count += _flush(current_chunk, current_header)

    cursor.execute("""
        INSERT INTO system_settings(key, value) VALUES('manual_fts_last_chunk_count', ?)
        ON CONFLICT(key) DO UPDATE SET value=excluded.value, updated_at=CURRENT_TIMESTAMP
    """, (str(cursor.execute("SELECT COUNT(*) FROM manual_chunks").fetchone()[0]),))

    conn.commit()
    conn.close()
    logger.info(f"Ingested {chunk_count} chunks from text manual {filename} ({len(pages)} page block(s)).")
    return chunk_count


def ingest_pdf_manuals(filepath: str = None) -> int:
    """
    Ingest PDF manuals into manual_chunks, keyed per source file.

    Called with a specific filepath (a fresh admin upload), only that one
    file's chunks are touched — every other manual's chunks stay exactly as
    they were. Called with no argument (Force Sync / full resync), every
    *.pdf currently present in ROOT_DIR is (re)processed. This is the fix
    for uploads silently not "taking": previously this always looked for
    exactly two hardcoded filenames, so a newly uploaded manual under any
    other name was never indexed at all.
    """
    conn = get_db_connection()
    cursor = conn.cursor()

    total_chunks = 0
    if filepath:
        total_chunks = _ingest_single_pdf(cursor, filepath)
    else:
        for fname in sorted(os.listdir(ROOT_DIR)):
            if fname.lower().endswith(".pdf"):
                total_chunks += _ingest_single_pdf(cursor, os.path.join(ROOT_DIR, fname))

    cursor.execute("""
        INSERT OR REPLACE INTO system_settings (key, value, updated_at)
        VALUES ('last_sync_time', ?, CURRENT_TIMESTAMP)
    """, (datetime.now().strftime("%Y-%m-%d %H:%M:%S"),))

    # Update manual_fts_last_chunk_count setting so DB init doesn't need to rebuild FTS immediately
    cursor.execute("""
        INSERT INTO system_settings(key, value) VALUES('manual_fts_last_chunk_count', ?)
        ON CONFLICT(key) DO UPDATE SET value=excluded.value, updated_at=CURRENT_TIMESTAMP
    """, (str(total_chunks),))

    conn.commit()
    conn.close()
    logger.info(f"Ingested {total_chunks} PDF chunks successfully.")
    return total_chunks


def run_full_ingestion():
    """
    Re-ingest every Excel and PDF source currently present.

    The resolved source filenames are reported back alongside the row counts.
    A spreadsheet that cannot be found is the one failure mode that is
    otherwise invisible -- it returns zero rows without touching the existing
    data, so the app carries on looking healthy while a sync quietly did
    nothing. Naming the file that was actually used makes that obvious.
    """
    init_database()

    sources = {
        "users": find_latest_source("users"),
        "hlb_allocations": find_latest_source("hlb_allocation"),
        "hlb_descriptions": find_latest_source("hlb_description"),
    }

    u_count = ingest_all_users()
    h_count = ingest_hlb_allocation()
    d_count = ingest_hlb_description()
    p_count = ingest_pdf_manuals()

    resolved = {k: (os.path.basename(v) if v else None) for k, v in sources.items()}
    missing = [k for k, v in sources.items() if not v]

    logger.info(
        f"Full Ingestion complete: {u_count} users, {h_count} HLB allocations, "
        f"{d_count} HLB descriptions, {p_count} PDF chunks. Sources: {resolved}"
    )
    if missing:
        logger.warning(f"No source spreadsheet found for: {', '.join(missing)}")

    return {
        "users_count": u_count,
        "hlb_allocations_count": h_count,
        "hlb_descriptions_count": d_count,
        "pdf_chunks_count": p_count,
        "sources_used": resolved,
        "sources_missing": missing,
        "status": "Completed" if not missing else "Completed with missing sources"
    }

if __name__ == "__main__":
    run_full_ingestion()
