"""
End-to-end check of the Field Attendance feature against a running server.

Run the app first (PORT=8099 python -m backend.main), then:
    python tests/test_attendance_flow.py

It exercises the rules that matter most for the census register:
  * a resubmission on the same day updates the ONE existing row, never adds one
  * two different people on the same day get two separate rows
  * a missing location, a bad mobile number, or a missing first photo is refused
  * approval deletes the photo file from disk and locks the record
  * rejection keeps the photo and lets the user resubmit (back to PENDING)
  * the admin routes reject non-admin callers
  * the Excel export contains one row per record
"""

import io
import os
import sys
import requests

BASE = os.environ.get("ATTENDANCE_TEST_BASE", "http://localhost:8099")
ADMIN_USER = os.environ.get("ADMIN_USERNAME", "testadmin")
ADMIN_PASS = os.environ.get("ADMIN_PASSWORD", "testpass")

PHOTO_DIR = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "attendance_photos")

PASSED, FAILED = 0, 0


def check(label, condition, detail=""):
    global PASSED, FAILED
    if condition:
        PASSED += 1
        print(f"  PASS  {label}")
    else:
        FAILED += 1
        print(f"  FAIL  {label} {detail}")


def photo_bytes(tag: bytes = b"a") -> bytes:
    # A tiny but structurally valid JPEG is not needed — the backend only
    # checks extension/mimetype and size — so a small unique blob is enough.
    return b"\xff\xd8\xff\xe0" + tag * 512 + b"\xff\xd9"


def submit(mobile, name, position, block, lat=24.9, lon=93.0, acc=8.0, photo=True, tag=b"a", token=None):
    files = {"photo": ("shot.jpg", io.BytesIO(photo_bytes(tag)), "image/jpeg")} if photo else None
    data = {"mobile_number": mobile, "name": name, "position": position, "block_number": block}
    if lat is not None:
        data["latitude"] = lat
    if lon is not None:
        data["longitude"] = lon
    if acc is not None:
        data["accuracy_m"] = acc
    headers = {"Authorization": f"Bearer {token}"} if token else {}
    return requests.post(f"{BASE}/api/attendance/submit", data=data, files=files, headers=headers, timeout=30)


def count_photo_files():
    total = 0
    for root, _dirs, files in os.walk(PHOTO_DIR):
        total += len(files)
    return total


def main():
    print(f"\nAttendance flow test against {BASE}\n" + "-" * 58)

    # --- admin session -----------------------------------------------------
    res = requests.post(f"{BASE}/api/auth/admin-login",
                        json={"username": ADMIN_USER, "password": ADMIN_PASS}, timeout=30).json()
    if not res.get("success"):
        print("Could not log in as admin — set ADMIN_USERNAME/ADMIN_PASSWORD to match the running server.")
        sys.exit(1)
    admin = {"Authorization": f"Bearer {res['token']}"}

    guest = requests.post(f"{BASE}/api/auth/guest", timeout=30).json()
    guest_headers = {"Authorization": f"Bearer {guest['token']}"}

    mobile_a, mobile_b = "9876543210", "9812345678"

    # Start from a clean slate for these two test numbers.
    existing = requests.get(f"{BASE}/api/admin/attendance?limit=500", headers=admin, timeout=30).json()
    for rec in existing.get("records", []):
        if rec["mobile_number"] in (mobile_a, mobile_b):
            requests.delete(f"{BASE}/api/admin/attendance/{rec['id']}", headers=admin, timeout=30)

    # --- validation --------------------------------------------------------
    print("\nValidation")
    check("short mobile number is refused",
          submit("12345", "Test User", "Enumerator", "0101").status_code == 400)
    check("missing location is refused",
          submit(mobile_a, "Test User", "Enumerator", "0101", lat=None, lon=None).status_code == 400)
    check("unknown position is refused",
          submit(mobile_a, "Test User", "Inspector", "0101").status_code == 400)
    check("blank block number is refused",
          submit(mobile_a, "Test User", "Enumerator", "  ").status_code == 400)
    check("first submission without a photo is refused",
          submit(mobile_a, "Test User", "Enumerator", "0101", photo=False).status_code == 400)

    # --- first submission --------------------------------------------------
    print("\nSubmission and deduplication")
    body = submit(mobile_a, " Ahmed   Ali ", "enumerator", "0142", tag=b"a").json()
    check("first submission succeeds", body.get("success") and body.get("created"), body.get("error", ""))
    rec_a = body["record"]
    check("name is whitespace-normalised", rec_a["name"] == "Ahmed Ali", rec_a["name"])
    check("position is normalised to title case", rec_a["position"] == "Enumerator", rec_a["position"])
    check("new record starts PENDING", rec_a["status"] == "PENDING")
    check("photo is attached", rec_a["has_photo"] is True)

    # --- resubmission updates the same row ---------------------------------
    body2 = submit(mobile_a, "Ahmed Ali", "Enumerator", "0155", lat=25.1, lon=93.2, tag=b"b").json()
    check("resubmission succeeds", body2.get("success") is True)
    check("resubmission does NOT create a new record", body2.get("created") is False)
    check("resubmission reuses the same row id", body2["record"]["id"] == rec_a["id"])
    check("edited block number is saved", body2["record"]["block_number"] == "0155")
    check("submission_count increments", body2["record"]["submission_count"] == 2)

    listing = requests.get(f"{BASE}/api/admin/attendance?q={mobile_a}", headers=admin, timeout=30).json()
    check("register holds exactly one row for this person today",
          len(listing["records"]) == 1, f"got {len(listing['records'])}")

    # --- a second person is a separate row ---------------------------------
    body3 = submit(mobile_b, "Rita Das", "Supervisor", "07", tag=b"c").json()
    check("a different mobile number creates its own row",
          body3.get("created") is True and body3["record"]["id"] != rec_a["id"])

    # --- carry-forward prefill --------------------------------------------
    print("\nLookup and carry-forward")
    look = requests.get(f"{BASE}/api/attendance/lookup?mobile={mobile_a}", timeout=30).json()
    check("lookup returns today's record", look["record"] and look["record"]["id"] == rec_a["id"])
    check("lookup returns a carry-forward profile", look["profile"]["name"] == "Ahmed Ali")
    check("lookup rejects an invalid number",
          requests.get(f"{BASE}/api/attendance/lookup?mobile=123", timeout=30).status_code == 400)

    # --- admin gating ------------------------------------------------------
    print("\nAdmin gating")
    check("register is not readable without a token",
          requests.get(f"{BASE}/api/admin/attendance", timeout=30).status_code == 403)
    check("register is not readable by a guest",
          requests.get(f"{BASE}/api/admin/attendance", headers=guest_headers, timeout=30).status_code == 403)
    check("photo is not readable by a guest",
          requests.get(f"{BASE}/api/admin/attendance/{rec_a['id']}/photo",
                       headers=guest_headers, timeout=30).status_code == 403)
    check("approve is not callable by a guest",
          requests.post(f"{BASE}/api/admin/attendance/{rec_a['id']}/approve",
                        headers=guest_headers, timeout=30).status_code == 403)

    # --- photo serving -----------------------------------------------------
    print("\nPhoto handling")
    photo_res = requests.get(f"{BASE}/api/admin/attendance/{rec_a['id']}/photo", headers=admin, timeout=30)
    check("admin can fetch the photo", photo_res.status_code == 200 and len(photo_res.content) > 0)

    before = count_photo_files()

    # --- rejection keeps the photo ----------------------------------------
    print("\nRejection")
    check("rejection without a reason is refused",
          requests.post(f"{BASE}/api/admin/attendance/{rec_a['id']}/reject",
                        json={"reason": ""}, headers=admin, timeout=30).status_code == 400)
    rej = requests.post(f"{BASE}/api/admin/attendance/{rec_a['id']}/reject",
                        json={"reason": "Photo is blurred, please retake."},
                        headers=admin, timeout=30).json()
    check("rejection succeeds", rej.get("success") is True)
    check("record is REJECTED", rej["record"]["status"] == "REJECTED")
    check("rejection keeps the photo", rej["record"]["has_photo"] is True)
    check("photo file is still on disk", count_photo_files() == before)

    # --- user corrects and resubmits --------------------------------------
    fixed = submit(mobile_a, "Ahmed Ali", "Enumerator", "0155", tag=b"d").json()
    check("a rejected entry can be corrected", fixed.get("success") is True)
    check("correcting returns it to PENDING", fixed["record"]["status"] == "PENDING")
    check("rejection reason is cleared", not fixed["record"]["reject_reason"])
    check("replacing the photo does not leave the old file behind",
          count_photo_files() == before, f"{count_photo_files()} vs {before}")

    # --- approval deletes the photo and locks the row ----------------------
    print("\nApproval")
    app_res = requests.post(f"{BASE}/api/admin/attendance/{rec_a['id']}/approve",
                            headers=admin, timeout=30).json()
    check("approval succeeds", app_res.get("success") is True)
    check("record is APPROVED", app_res["record"]["status"] == "APPROVED")
    check("record reports its photo as deleted", app_res["record"]["photo_deleted"] is True)
    check("record no longer exposes a photo", app_res["record"]["has_photo"] is False)
    check("photo file is gone from disk", count_photo_files() == before - 1,
          f"{count_photo_files()} vs expected {before - 1}")
    check("photo endpoint now 404s",
          requests.get(f"{BASE}/api/admin/attendance/{rec_a['id']}/photo",
                       headers=admin, timeout=30).status_code == 404)

    locked = submit(mobile_a, "Ahmed Ali", "Enumerator", "9999", tag=b"e")
    check("an approved entry cannot be edited", locked.status_code == 409)
    check("the locked response says so", locked.json().get("locked") is True)

    after_lock = requests.get(f"{BASE}/api/attendance/lookup?mobile={mobile_a}", timeout=30).json()
    check("the approved block number was not overwritten",
          after_lock["record"]["block_number"] == "0155", after_lock["record"]["block_number"])
    check("editable flag is false once approved", after_lock["record"]["editable"] is False)

    # --- filters -----------------------------------------------------------
    print("\nFilters and export")
    approved_only = requests.get(f"{BASE}/api/admin/attendance?status=APPROVED", headers=admin, timeout=30).json()
    check("status filter returns only approved rows",
          all(r["status"] == "APPROVED" for r in approved_only["records"]) and approved_only["total"] >= 1)
    sup_only = requests.get(f"{BASE}/api/admin/attendance?position=Supervisor", headers=admin, timeout=30).json()
    check("position filter works",
          all(r["position"] == "Supervisor" for r in sup_only["records"]) and sup_only["total"] >= 1)
    empty = requests.get(f"{BASE}/api/admin/attendance?date_from=2000-01-01&date_to=2000-01-02",
                         headers=admin, timeout=30).json()
    check("date range filter excludes out-of-range rows", empty["total"] == 0)

    # --- excel export ------------------------------------------------------
    exp = requests.get(f"{BASE}/api/admin/attendance/export", headers=admin, timeout=60)
    check("export returns an xlsx", exp.status_code == 200 and
          exp.headers.get("Content-Type", "").startswith("application/vnd.openxml"))
    check("export is offered as a download", "attachment" in exp.headers.get("Content-Disposition", ""))
    check("export is not downloadable by a guest",
          requests.get(f"{BASE}/api/admin/attendance/export", headers=guest_headers, timeout=30).status_code == 403)

    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(exp.content))
    ws = wb.active
    check("workbook has a single Attendance sheet",
          wb.sheetnames == ["Attendance"], str(wb.sheetnames))
    headers = [c.value for c in ws[1]]
    check("workbook header includes the required columns",
          all(h in headers for h in ["Date", "Name", "Mobile Number", "Position",
                                     "HLB / Circle No", "Latitude", "Longitude", "Status", "Photo"]),
          str(headers))
    total_rows = requests.get(f"{BASE}/api/admin/attendance?limit=500", headers=admin, timeout=30).json()["total"]
    check("one worksheet row per register row",
          ws.max_row == total_rows + 1, f"{ws.max_row - 1} rows vs {total_rows} records")
    names = {ws.cell(row=r, column=2).value for r in range(2, ws.max_row + 1)}
    check("multiple users share one sheet", {"Ahmed Ali", "Rita Das"} <= names, str(names))
    photo_col = headers.index("Photo") + 1
    photo_notes = {ws.cell(row=r, column=photo_col).value for r in range(2, ws.max_row + 1)}
    check("approved rows record that the photo was deleted",
          "Deleted after approval" in photo_notes, str(photo_notes))

    filtered = requests.get(f"{BASE}/api/admin/attendance/export?position=Supervisor", headers=admin, timeout=60)
    fws = load_workbook(io.BytesIO(filtered.content)).active
    check("filtered export honours the filter",
          all(fws.cell(row=r, column=4).value == "Supervisor" for r in range(2, fws.max_row + 1)))

    # --- housekeeping ------------------------------------------------------
    print("\nHousekeeping")
    purge = requests.post(f"{BASE}/api/admin/attendance/purge-photos", headers=admin, timeout=30).json()
    check("orphan purge runs", purge.get("success") is True)
    check("nothing was orphaned", purge.get("removed") == 0, str(purge.get("removed")))

    print("-" * 58)
    print(f"{PASSED} passed, {FAILED} failed\n")
    sys.exit(1 if FAILED else 0)


if __name__ == "__main__":
    main()
